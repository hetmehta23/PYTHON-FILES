import openpyxl
import calendar
import datetime
    

file=openpyxl.load_workbook(filename="Trade_act.xlsx",data_only=True)
fs= file.get_sheet_by_name("Sheet1")
fs1=file.create_sheet("Final")
fs1= file.get_sheet_by_name("Final")
fs1.cell(row=1,column=1).value="RowID"
fs1.cell(row=1,column=2).value="Date"
fs1.cell(row=1,column=3).value="Day of the week"
fs1.cell(row=1,column=4).value="Week Ending Friday Date"
fs1.cell(row=1,column=5).value="Segment"
fs1.cell(row=1,column=6).value="Trade-Act/Day"


def value_to_date(value):
    date=str(value)
    d=datetime.datetime.strptime(date,'%Y-%m-%d %H:%M:%S')
    return d
    
q=2
for row_id in range(2,fs.max_row+1):
    id_no=fs.cell(row=row_id,column=1).value
    f_date=fs.cell(row=row_id,column=2).value
    s_date=value_to_date(f_date)
    t_date=fs.cell(row=row_id,column=4).value
    e_date=value_to_date(t_date)
    no_of_days=fs.cell(row=row_id,column=5).value
    segment=fs.cell(row=row_id,column=6).value
    trade=int(fs.cell(row=row_id,column=7).value)

    l=1
    while (l<=no_of_days):
        fs1.cell(row=q,column=1).value=id_no
        fs1.cell(row=q,column=2).value=s_date.strftime('%m/%d/%Y')
        
        week_day=int(s_date.strftime('%w'))
        if(week_day==6):
            week_day=1
        else:
            week_day=week_day+2
            
        fs1.cell(row=q,column=3).value=week_day

        if(week_day==1):
            end_date=s_date+datetime.timedelta(days=6)
        elif(week_day==2):
            end_date=s_date+datetime.timedelta(days=5)
        elif(week_day==3):
            end_date=s_date+datetime.timedelta(days=4)
        elif(week_day==4):
            end_date=s_date+datetime.timedelta(days=3)
        elif(week_day==5):
            end_date=s_date+datetime.timedelta(days=2)
        elif(week_day==6):
            end_date=s_date+datetime.timedelta(days=1)
        elif(week_day==7):
            end_date=s_date
        fs1.cell(row=q,column=4).value=end_date.strftime('%m/%d/%Y')        
        s_date=s_date+datetime.timedelta(days=1)    
        fs1.cell(row=q,column=5).value=segment
        fs1.cell(row=q,column=6).value=trade
        l=l+1
        q=q+1


file.save('final.xlsx')

